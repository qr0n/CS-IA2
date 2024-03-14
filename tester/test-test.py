from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import pyautogui
import pygetwindow as gw
import time

workbook_path = 'C:/Users/bhave/OneDrive/Documents/test.xlsx'
workbook = load_workbook(workbook_path)
alphabet_dict = {i: chr(i + ord('A') - 1) for i in range(1, 27)}

sheet = workbook['Sheet1']
left, top, width, height = 100, 100, 500, 500

class ExcelFileManagement:
    @staticmethod
    def add_value(cell_id, text):
        try:
            sheet[cell_id] = text
            workbook.save(workbook_path)
            return "Done"
        except Exception as E:
                return E
        
    @staticmethod
    def add_image_to_cell(cell_id, image_path, width, height):
        try:
            img = Image(image_path)
            img.width = width
            img.height = height
            sheet.add_image(img, cell_id)
            workbook.save(workbook_path)
            return "Done"
        except Exception as E:
            return E
            
    @staticmethod
    def populate_table():
        headers = ["Test number", "Test Type", "Variables", "Module", "Input", "Expected result", "Actual result", "Screenshot"]
    # Calculate the starting cell for populating data
        start_row = sheet.max_row + 5
        start_cell = f"{alphabet_dict[1]}{start_row}"

        for i, header in enumerate(headers, start=1):
            sheet.cell(row=i + start_row - 1, column=1, value=header)

        print(f"Starting from cell: {start_cell}")
        workbook.save(workbook_path)
        
    @staticmethod
    def add_results(module, input_data, expected):
        try:
        # Find the next available row
            row = sheet.max_row + 1
        # Set the column index to 2 (column B)
            col = 2
        # Get the cell IDs for each row in column B
            module_cell = f"{alphabet_dict[col]}{row}"
            input_cell = f"{alphabet_dict[col]}{row + 1}"
            expected_cell = f"{alphabet_dict[col]}{row + 2}"

        # Set values in the corresponding cells
            sheet[module_cell] = module
            sheet[input_cell] = input_data
            sheet[expected_cell] = expected

        # Save the workbook
            workbook.save(workbook_path)
            return "Results added successfully"
        except Exception as e:
            return str(e)



class ScreenshotManagement:
    class Helper:
        def focus_window_by_title(window_title):
            try:
                window = gw.getWindowsWithTitle(window_title)[0]
                window.activate()
                return True
            except IndexError:
                print(f"Window with title '{window_title}' not found.")
                return False
            
    @staticmethod
    def take_screenshot(path_to_save):
        ScreenshotManagement.Helper.focus_window_by_title("Windows PowerShell")
        time.sleep(1)
        screenshot = pyautogui.screenshot(region=(left, top, width, height))
        screenshot.save(path_to_save)
    
# ScreenshotManagement.take_screenshot("C:/Users/bhave/OneDrive/Desktop/test.png")
# ExcelFileManagement.populate_table()
ExcelFileManagement.add_results("main", "hello data", "nothing yet")

""" TODO 
- use 2 functions, one for headers, one for data ( a little cancer but its easier than what was being done before )
"""